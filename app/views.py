from django.shortcuts import render
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
import time
import re

DURATION_SECONDS = 10 * 60  # 10 minutes


def _safe_float(x):
    try:
        return float(x)
    except Exception:
        return None


def _normalize_formula(s: str) -> str:
    # Retire espaces, majuscules, uniformise séparateurs FR/EN
    s = (s or "").replace(" ", "").upper()
    s = s.replace(";", ",")
    return s


def _mentions_all_cells(norm: str, cells: list[str]) -> bool:
    # Vérifie que chaque cellule apparaît dans la formule (B3, B4, ...)
    return all(c in norm for c in cells)


def _looks_like_sum_solution(norm: str, expected_cells: list[str]) -> tuple[bool, str]:
    """
    Retourne (ok, reason) si la formule ressemble à une solution valide.
    Solutions acceptées (tolérant) :
      - SOMME(B3:B7) / SUM(B3:B7)
      - SOMME(B3;B4;B5;B6;B7) / SUM(B3,B4,B5,B6,B7)
      - B3+B4+B5+B6+B7
    """
    range_ok = "B3:B7" in norm
    list_ok = _mentions_all_cells(norm, expected_cells) and ("," in norm or "(" in norm)
    add_ok = _mentions_all_cells(norm, expected_cells) and "+" in norm

    if range_ok:
        return True, "Plage B3:B7 détectée"
    if list_ok:
        return True, "Liste B3..B7 détectée"
    if add_ok:
        return True, "Addition B3..B7 détectée"
    return False, "Références B3..B7 non détectées"


@csrf_exempt
def test_upload(request):
    now = int(time.time())

    # Timer via cookie (plus fiable que session sur Render free)
    start_ts = request.COOKIES.get("test_start_ts")
    if start_ts is None:
        start_ts = str(now)

    elapsed = now - int(start_ts)
    remaining = max(0, DURATION_SECONDS - elapsed)

    # GET : page test + timer
    if request.method == "GET":
        resp = render(request, "test_upload.html", {"remaining_seconds": remaining})
        if "test_start_ts" not in request.COOKIES:
            resp.set_cookie("test_start_ts", start_ts, max_age=DURATION_SECONDS, samesite="Lax")
        return resp

    # POST : si temps écoulé -> résultat
    if remaining <= 0:
        return render(
            request,
            "result.html",
            {
                "verdict": "⏰ Temps écoulé",
                "score": 0,
                "remaining_seconds": 0,
                "feedback": ["Upload refusé : le temps est écoulé."],
            },
            status=403,
        )

    f = request.FILES.get("file")
    if not f:
        return render(
            request,
            "result.html",
            {
                "verdict": "❌ Fichier manquant",
                "score": 0,
                "remaining_seconds": remaining,
                "feedback": ["Aucun fichier n’a été reçu."],
            },
            status=400,
        )

    # Charger Excel
    try:
        wb = openpyxl.load_workbook(f, data_only=False)
    except Exception as e:
        return render(
            request,
            "result.html",
            {
                "verdict": "❌ Fichier invalide",
                "score": 0,
                "remaining_seconds": remaining,
                "feedback": [f"Le fichier n’est pas lisible (.xlsx attendu). Détail : {e}"],
            },
            status=400,
        )

    ws = wb.active
    formula = ws["B2"].value

    # Barème /20
    score = 0
    feedback = []

    # +10 si B2 est une formule (pas une valeur en dur)
    if isinstance(formula, str) and formula.startswith("="):
        score += 10
        feedback.append("✅ Formule détectée en B2 (+10).")
    else:
        feedback.append(f"❌ B2 n'est pas une formule (valeur trouvée : {formula}) (+0).")
        return render(
            request,
            "result.html",
            {
                "verdict": "🔴 À revoir",
                "score": score,
                "remaining_seconds": remaining,
                "feedback": feedback,
            },
            status=200,
        )

    # Multi-solutions tolérées
    norm = _normalize_formula(formula)
    expected_cells = ["B3", "B4", "B5", "B6", "B7"]

    # +5 si la formule référence correctement B3..B7 (plage, liste, addition)
    ok_refs, reason = _looks_like_sum_solution(norm, expected_cells)
    if ok_refs:
        score += 5
        feedback.append(f"✅ Références OK : {reason} (+5).")
    else:
        feedback.append(f"⚠️ Références attendues B3..B7 non trouvées : {formula} (+0).")

    # Calcul de la somme attendue à partir des valeurs (B3..B7)
    values = []
    for r in range(3, 8):
        v = _safe_float(ws[f"B{r}"].value)
        if v is not None:
            values.append(v)

    expected_sum = sum(values) if values else 0.0
    feedback.append(f"ℹ️ Somme attendue (B3:B7) = {expected_sum:g}")

    # +5 si la formule est cohérente : SUM/SOMME ou addition explicite
    is_sum_function = ("SUM(" in norm) or ("SOMME(" in norm)
    is_explicit_add = ("+" in norm) and _mentions_all_cells(norm, expected_cells)

    if values and ok_refs and (is_sum_function or is_explicit_add):
        score += 5
        feedback.append("✅ Solution de somme reconnue (SUM/SOMME ou addition) (+5).")
    else:
        feedback.append("⚠️ Cohérence non validée (formule non reconnue, refs manquantes ou données vides) (+0).")

    # Verdict
    if score == 20:
        verdict = "🎉 Parfait !"
    elif score >= 15:
        verdict = "✅ Très bien"
    elif score >= 10:
        verdict = "🟠 Correct"
    else:
        verdict = "🔴 À revoir"

    return render(
        request,
        "result.html",
        {
            "verdict": verdict,
            "score": score,
            "remaining_seconds": remaining,
            "feedback": feedback,
        },
        status=200,
    )
